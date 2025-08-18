from django.db import models
from django.urls import reverse
from django.utils import timezone
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime  # Correct import
from django.apps import apps


from django.contrib.auth.models import AbstractBaseUser, BaseUserManager
from django.contrib.auth.models import PermissionsMixin

from django.conf import settings

from django.db.models import Sum


class UserManager(BaseUserManager):

    def create_user(self, email, password=None, **extra_fields):
        # Creates and save a new user
        if not email:
            raise ValueError('Users must have an email address!')

        user = self.model(email=self.normalize_email(email), **extra_fields)
        user.set_password(password)
        user.save(using=self._db)

        return user

    def create_superuser(self, email, password):
        # Creates and save a new superuser
        user = self.create_user(email, password)
        user.is_staff = True
        user.is_superuser = True
        user.save(using=self._db)
        return user


class User(PermissionsMixin, AbstractBaseUser):
    # Custom user model supports email instead of username
    email = models.EmailField(max_length=255, unique=True)
    name = models.CharField(max_length=255)
    is_active = models.BooleanField(default=True)
    is_staff = models.BooleanField(default=False)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    objects = UserManager()

    USERNAME_FIELD = 'email'


ACCOUNT_TYPE = (
    ('accrual', 'ACCRUAL'),
    ('cash', 'CASH'),
)


class Company(models.Model):
    user = models.ForeignKey('User', null=True, on_delete=models.SET_NULL)
    title = models.CharField(max_length=200)
    parent_company = models.ForeignKey(
        'Company', related_name='Parent', on_delete=models.CASCADE, null=True, blank=True)
    description = models.TextField(blank=True, null=True)
    owners = models.ManyToManyField('Owner')
    address = models.ForeignKey(
        'Address', on_delete=models.CASCADE, blank=True, null=True)
    account_type = models.CharField(
        max_length=20,  default='accrual', choices=ACCOUNT_TYPE)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    """
    1 IK    -
    2 Fig   IK
    3 Biza  IK
    """

    def __str__(self):
        return self.title

    def get_absolute_url(self):
        return reverse('company-detail', kwargs={'pk': self.pk})


class Address(models.Model):
    addr_line1 = models.CharField(max_length=400, blank=True, null=True)
    city = models.CharField(max_length=200, blank=True, null=True)
    state = models.CharField(max_length=200, blank=True, null=True)
    country = models.CharField(max_length=200, blank=True, null=True)
    zipcode = models.CharField(max_length=10, blank=True, null=True)
    phone = models.CharField(max_length=200, blank=True, null=True)
    email = models.EmailField(max_length=200, blank=True, null=True)

    #latitude = models.FloatField(null=True, blank=True);
    #longitude = models.FloatField(null=True, blank=True);

    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)


class Owner(models.Model):
    name = models.CharField(max_length=200)
    share = models.DecimalField(max_digits=2, decimal_places=2)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return self.name

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class Account(models.Model):
    pass
    """
    1 Balance Sheet
    """


class AccountTemplate(models.Model):
    title = models.CharField(max_length=200)
    description = models.TextField(blank=True, null=True)
    dr = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    cr = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    balance = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'{self.title} - dr: {self.dr}'


ACCOUNT_SECTIONS = (
    ("Assets", "ASSETS"),
    ("Liabilities", "LIABILITIES"),
    ("Equity", "EQUITY"),
    ("Revenue", "REVENUE"),
    ("Expense", "EXPENSE"),
)


class AccountMain(AccountTemplate):
    company = models.ForeignKey('Company', on_delete=models.CASCADE)
    account_section = models.CharField(max_length=30, choices=ACCOUNT_SECTIONS)

    def __str__(self):
        return f'{self.title} ({self.account_section})'

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class AccountSub(AccountTemplate):
    main = models.ForeignKey('AccountMain', on_delete=models.CASCADE)

    def __str__(self):
        return f'{self.title} ({self.main})'

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class Item(AccountTemplate):
    account_sub = models.ForeignKey('AccountSub', on_delete=models.CASCADE)
    item_depretiation = models.ForeignKey('DSP', on_delete=models.CASCADE)

    def __str__(self):
        return f'{self.title} - {self.account_sub}'

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


# TANSACTION_CATEGORY = (
#     ("Purchase", "ASSETS"),
#     ("Sale", "LIABILITIES"),
#     ("Payment", "EQUITY"),
#     ("Receive", "REVENUE"),
# )

"""
MODELS: CUSTOMER,   VENDOR, SUPPLIER
"""


class CVbase(models.Model):
    name = models.CharField(max_length=100, blank=True, null=True)

    address = models.ForeignKey(
        'Address', on_delete=models.CASCADE, blank=True, null=True
    )
    phone = models.CharField(max_length=30, blank=True, null=True)
    balance = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    note = models.TextField(blank=True, null=True)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)


class Customer(CVbase):
    account_type = models.ForeignKey(
        'AccountSub', on_delete=models.CASCADE, null=True, blank=True)

    def __str__(self):
        return self.name

    @property
    def get_account_sub(self):
        return self.account_type

    def save(self, *args, **kwargs):
        if not self.account_type:
            a_sub = AccountSub.objects.get(title='Account Receivable')
            self.account_type = a_sub
            print(self.account_type)
            # self.total = self.unit_price * self.quantity
        super(Customer, self).save(*args, **kwargs)

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class Vendor(CVbase):
    account_type = models.ForeignKey(
        'AccountSub', on_delete=models.CASCADE, null=True, blank=True)

    def __str__(self):
        return self.name

    @property
    def get_account_sub(self):
        return self.account_type

    def save(self, *args, **kwargs):
        if not self.account_type:
            a_sub = AccountSub.objects.get(title='Account Payable')
            self.account_type = a_sub
            print(self.account_type)
            # self.total = self.unit_price * self.quantity
        super(Vendor, self).save(*args, **kwargs)

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


""" END """


# class TransactionType(models.Model):

#     debit_account = models.ForeignKey('Company', on_delete=models.CASCADE)
#     credit_account = models.CharField(max_length=30, blank=True, null=True)
#     # amount = models.IntegerField(default=0)
#     description = models.TextField(blank=True, null=True)
# credit_account

"""
    Purchase
    Sale
    Payment
    Receive
    """


class Purchase(models.Model):
    department = models.ForeignKey(
        'Company', 
        on_delete=models.CASCADE,
        default=1
        )
    vendor = models.ForeignKey('Vendor', on_delete=models.CASCADE)
    purchase_invoice = models.ForeignKey(
        'PurchaseInvoice', on_delete=models.CASCADE, null=True, blank=True)
    # amount = models.IntegerField(default=0)
    description = models.TextField(blank=True, null=True)
    # credit_account
    item = models.ForeignKey('Item', on_delete=models.CASCADE)
    quantity = models.PositiveIntegerField(default=1)
    unit_price = models.DecimalField(
        max_digits=9, decimal_places=2, default=0.00)
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    note = models.TextField(null=True, blank=True)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    # def save(self):
    #     return self.unit_price * self.quantity
    def save(self, *args, **kwargs):
        self.total = self.unit_price * self.quantity
        super(Purchase, self).save(*args, **kwargs)

    def __str__(self):
        return f'{self.item} - {self.total}'

    @property
    def get_account_sub(self):
        return self.item.account_sub

    def get_total_list_price(self):
        total = self.objects.all().aggregate(Sum('total'))
        # tot Purchase.objects.filter(
        #     credit_account=vendor).aggregate(Sum('total'))
        return total

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class Sale(models.Model):

    department = models.ForeignKey(
        'Company', on_delete=models.CASCADE, default=1)
    customer = models.ForeignKey('Customer', on_delete=models.CASCADE)
    invoice = models.ForeignKey(
        'Invoice', on_delete=models.CASCADE, null=True, blank=True)
    description = models.TextField(blank=True, null=True)
    item = models.ForeignKey('Item', on_delete=models.CASCADE)
    quantity = models.PositiveIntegerField(default=1)
    unit_price = models.DecimalField(
        max_digits=9, decimal_places=2, default=0.00)
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    balance = models.DecimalField(
        max_digits=10, decimal_places=3, default=0.00)
    status = models.CharField(max_length=30, default='unpaid')
    note = models.TextField(null=True, blank=True)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def save(self, *args, **kwargs):
        self.total = self.unit_price * self.quantity
        self.customer = self.invoice.customer
        super(Sale, self).save(*args, **kwargs)

    def __str__(self):
        return f'{self.pk} | {self.item} - {self.total} to {self.customer}'

    @property
    def get_account_sub(self):
        return self.item.account_sub

    @property
    def get_total_list_price(self, items):
        total = self.objects.all().aggregate(Sum('total'))
        # tot Purchase.objects.filter(
        #     credit_account=vendor).aggregate(Sum('total'))
        return total


    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})
PAYMENT_METHOD = (
    ('cash', 'CASH'),
    ('Credit Card', 'CREDIT CARD'),
    ('Cheque', 'CHEQUE')
)


class Payment(models.Model):

    from_account = models.ForeignKey(
        'Company', on_delete=models.CASCADE, default=1)
    vendor = models.ForeignKey('Vendor', on_delete=models.CASCADE)
    purchase_invoice = models.ForeignKey(
        'PurchaseInvoice', on_delete=models.CASCADE, null=True, blank=True)
    payment_method = models.CharField(
        max_length=20, choices=PAYMENT_METHOD, default='cash')
    description = models.TextField(null=True, blank=True)
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'{self.total} to {self.vendor}'

    @property
    def get_account_sub(self):
        return self.vendor.account_type

    def get_total_list_price(self):
        total = self.objects.all().aggregate(Sum('amount'))
        print(total)
        return total

    # def get_absolute_url(self):
    #     return reverse('company-detail', kwargs={'pk': self.pk})


class Receive(models.Model):

    to_account = models.ForeignKey(
        'Company', on_delete=models.CASCADE, default=1)
    from_account = models.ForeignKey('Customer', on_delete=models.CASCADE)
    invoice = models.ForeignKey(
        'Invoice', on_delete=models.CASCADE,  null=True, blank=True)
    # invoice = models.CharField(max_length=60, null=True, blank=True)
    payment_method = models.CharField(
        max_length=20, choices=PAYMENT_METHOD, default='cash')
    description = models.TextField(null=True, blank=True)
    amount = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'{self.pk} - {self.amount} from {self.from_account}'

    @property
    def get_account_sub(self):
        return self.from_account.account_type

    def get_absolute_url(self):
        return reverse('sales_receive_detail', kwargs={'pk': self.pk})

    def get_update_url(self):
        return reverse('sales_receive_update', args=(self.pk,))
# REPORTS


class Expense(models.Model):
    pass


class Revenue(models.Model):
    pass


INVOICE_STATUS = (
    ('pending', 'PENDING'),
    ('invoked', 'INVOKED'),
    ('paid', 'PAID')
)


class Invoice(models.Model):
    account = models.ForeignKey('Company', on_delete=models.CASCADE, default=1)
    customer = models.ForeignKey('Customer', on_delete=models.CASCADE)

    # sale_query = models.QuerySet(Sale.objects.filter(customer=customer))
    # total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)

    note = models.TextField(null=True, blank=True)
    due_date = models.DateTimeField(null=True, blank=True)
    payment_term = models.CharField(
        max_length=60, choices=PAYMENT_METHOD, default='cash'
    )
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    status = models.CharField(
        max_length=30, choices=INVOICE_STATUS, default='pending')
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    class Meta:
        ordering = ('-created',)

    def __init__(self, *args, **kwargs):
        super(Invoice, self).__init__(*args, **kwargs)

    def __str__(self):
        return f'{self.id} - {self.customer}'

    def get_absolute_url(self):
        return reverse('sales_invoice_detail', kwargs={'pk': self.pk})

    @property
    def get_sale_items(self):
        return InvoiceItem.objects.filter(invoice=self)

    @property
    def get_total_price(self):
        items = InvoiceItem.objects.filter(invoice=self)
        total = 0
        for item in items:
            total += item.total
        return total

    def get_update_url(self):
        return reverse('sales_invoice_update', args=(self.pk,))

    def pay(self):
        return reverse('sales_invoice_pay', args=(self.pk,))


class PurchaseInvoice(models.Model):

    # Fields
    # slug = extension_fields.AutoSlugField(populate_from='id', blank=True)
    account = models.ForeignKey('Company', on_delete=models.CASCADE, default=1)
    vendor = models.ForeignKey(
        'Vendor',
        on_delete=models.CASCADE,
        related_name="vendors"
    )
    payment_term = models.CharField(
        max_length=60, choices=PAYMENT_METHOD, default='cash'
    )

    # sale_query = models.QuerySet(Sale.objects.filter(customer=customer))
    # total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)

    note = models.TextField(null=True, blank=True)
    due_date = models.DateTimeField(null=True, blank=True)
    
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    status = models.CharField(
        max_length=30, choices=INVOICE_STATUS, default='pending')
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    class Meta:
        ordering = ('-created',)

    # def __unicode__(self):
    #     return u'%s' % self.slug

    def get_absolute_url(self):
        return reverse('purchase_invoice_detail', args=(self.pk,))

    def get_update_url(self):
        return reverse('purchase_invoice_update', args=(self.pk,))


class InvoiceItem(models.Model):
    invoice = models.ForeignKey('Invoice', on_delete=models.CASCADE)
    item = models.ForeignKey('Item', on_delete=models.CASCADE)
    description = models.CharField(max_length=200)

    quantity = models.PositiveIntegerField(default=1)
    unit_price = models.DecimalField(
        max_digits=9, decimal_places=2, default=0.00)
    total = models.DecimalField(max_digits=9, decimal_places=2, default=0.00)
    balance = models.DecimalField(
        max_digits=10, decimal_places=3, default=0.00)
    status = models.CharField(max_length=30, default='unpaid')
    note = models.TextField(null=True, blank=True)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'{self.invoice.id} - {self.item.title} ({self.pk})'

    def save(self, *args, **kwargs):
        self.total = self.unit_price * self.quantity
        super(InvoiceItem, self).save(*args, **kwargs)

    def get_absolute_url(self):
        return reverse(
            'sales_invoiceitem_detail',
            kwargs={'pk': self.pk}
        )

    # def calculate_total(self):
    #     return sum(item.total for item in self.sale_items.all())


DEPRECIATION_CHOICES = [
    (1, 'Depreciation'),
    (2, 'Prepaid'),
    (3, 'Supplies')
]

# D = Depreciation - S = Supplies - P = Prepaid


class DSP(models.Model):
    title = models.CharField(max_length=60, null=True, blank=True)
    depreciation_choices = models.PositiveIntegerField(
        choices=DEPRECIATION_CHOICES)
    # year_of_purchase = models.IntegerField()
    created = models.DateTimeField(auto_now_add=True)
    lifespan = models.PositiveIntegerField()

    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'{self.title} - {self.created}'

    # def save(self, *args, **kwargs):
    #     now = datetime.datetime.now()
    #     self.year_of_purchase = now.year
    #     super(Sale, self).save(*args, **kwargs)


class BalanceSheet(models.Model):
    """
    account_name = models.CharField(max_length=200)
    """
    pass


"""
For IncomeStatement:
1. If a payment transaction  is done for an expense the account of the expense
    will be recorded to the income statement
rent    $500    [operating expense]bv
"""

MODEL_CHOICES = (
    ('Purchase', 'PURCHASE'),
    ('Sale', 'SALE'),
    ('Payment', 'PAYMENT'),
    ('Receive', 'RECEIVE'),
)

from django.contrib.contenttypes.models import ContentType
from django.contrib.contenttypes.fields import GenericForeignKey

class Journal(models.Model):
    created = models.DateTimeField(auto_now_add=True)
    dr_account = models.ForeignKey(
        'AccountSub', on_delete=models.CASCADE, related_name='debit_account')
    cr_account = models.ForeignKey(
        'AccountSub', on_delete=models.CASCADE, related_name='credit_account')
    content_type = models.ForeignKey(ContentType, on_delete=models.CASCADE, null=True, blank=True)
    object_id = models.PositiveIntegerField(null=True, blank=True)
    content_object = GenericForeignKey('content_type', 'object_id')
    sender_model = models.CharField(max_length=10, null=True, blank=True)

    model_id = models.PositiveIntegerField(null=True, blank=True)

    description = models.TextField(blank=True, null=True)
    amount = models.DecimalField(max_digits=12, decimal_places=2, default=0.00)

    updated = models.DateTimeField(auto_now=True)

    def __str__(self):
        # return f'Dr: {self.dr_account}-{self.amount} --- Cr: {self.cr_account}-{self.amount}'
        return f'{self.created} - {self.dr_account}'

    @property
    def get_transaction(self):
        transaction_model = apps.get_model('core', self.sender_model)
        transaction = transaction_model.objects.get(id=self.model_id)
        return transaction

    @property
    def get_output(self):
        output = self.dr_account.main.account_section
        if output == 'Assets' or output == 'Liability' or output == 'Equity':
            print(self.dr_account.main.account_section)
            return 'balance'
        else:
            return 'income'

    @classmethod
    def export_to_excel(cls, queryset, filename=None):
        """
        Exports Journal queryset to Excel
        Args:
            queryset: Journal queryset to export
            filename: Custom filename (optional)
        Returns:
            Tuple of (file_data, filename) for further handling
        """
        import os
        from django.conf import settings
        EXPORT_DIR = os.path.join(settings.BASE_DIR, 'journal_exports')
        os.makedirs(EXPORT_DIR, exist_ok=True)
    
        filename = f"journal_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        if not filename:
            filename = f"journal_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Journal Entries"
        
        # Column headers
        headers = [
            'ID', 'Created Date', 'Debit Account', 'Credit Account', 
            'Sender Model', 'Model ID', 'Description', 'Amount', 'Updated Date'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data rows
        for row_num, journal in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=journal.id)
            ws.cell(row=row_num, column=2, value=journal.created.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=3, value=str(journal.dr_account))
            ws.cell(row=row_num, column=4, value=str(journal.cr_account))
            ws.cell(row=row_num, column=5, value=journal.sender_model or '')
            ws.cell(row=row_num, column=6, value=journal.model_id or '')
            ws.cell(row=row_num, column=7, value=journal.description or '')
            ws.cell(row=row_num, column=8, value=float(journal.amount))
            ws.cell(row=row_num, column=9, value=journal.updated.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        # Save to bytes buffer instead of HttpResponse
        buffer = BytesIO()
        wb.save(filepath)
        buffer.seek(0)
        
        return filepath

class IncomeStatement(models.Model):

    model = models.CharField(max_length=10, choices=MODEL_CHOICES)
    model_id = models.PositiveIntegerField()
    account = models.ForeignKey('AccountSub', on_delete=models.CASCADE)

    def __str__(self):
        return f'{self.model} - {self.model_id} ({self.account})'
